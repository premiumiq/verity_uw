"""Generate the HITL cheat-sheet Excel from seed_uw.py SUBMISSIONS.

Output: docs/guides/seed_cheatsheet.xlsx with one tab per submission
in two flavours:

  * <short_id>     — key/value view of every seeded field on the
                     submission row. Two columns, "Field" and
                     "Value". One row per field.

  * LH-<short_id>  — loss-history view as a flat table. Columns are
                     Policy Year / Claims / Incurred / Paid /
                     Reserves. Header row at the top names the
                     submission so each tab is usable on its own.

The Excel file is only consumed by humans during HITL review — it
holds the ground-truth values that the seeder used to ship to the
database before the minimal-seed change. The 16 workflow rows
(everything except 00000001 + 00000016) now leave most of those
fields NULL in the database, so this file is the operator's
reference for what to type in when the AI extractor misses.

Run from the repo root:

    .venv/bin/python -m docs.guides.build_seed_cheatsheet

(openpyxl is installed in .venv but intentionally not added to
requirements.txt; this script is operator tooling, not app code.)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from uw_demo.app.setup.seed_uw import SUBMISSIONS, SHOWCASE_IDS


# ── Output location ──────────────────────────────────────────
# Sit next to the showcase doc so all HITL operator tooling
# lives in one folder.

OUT_PATH = Path(__file__).parent / "seed_cheatsheet.xlsx"


# ── Field column ordering ────────────────────────────────────
# Match the visual order on the Submission Details tab so the
# cheat sheet reads top-to-bottom the same way the UI does.
# `id` and `showcase` are added at the very top for quick
# orientation; everything else mirrors field_layout.py grouping.

FIELD_ORDER = [
    # Identity / orientation
    ("id",                       "Submission UUID"),
    ("short_id",                 "Short ID"),
    ("showcase",                 "Showcase row"),
    # Company
    ("named_insured",            "Named Insured"),
    ("lob",                      "Line of Business"),
    ("fein",                     "FEIN"),
    ("entity_type",              "Entity Type"),
    ("state_of_incorporation",   "State of Incorporation"),
    ("sic_code",                 "SIC Code"),
    ("sic_description",          "Industry"),
    ("annual_revenue",           "Annual Revenue"),
    ("employee_count",           "Employee Count"),
    # Policy
    ("effective_date",           "Effective Date"),
    ("expiration_date",          "Expiration Date"),
    ("limits_requested",         "Limits Requested"),
    ("retention_requested",      "Retention"),
    # Prior coverage
    ("prior_carrier",            "Prior Carrier"),
    ("prior_premium",            "Prior Premium"),
    # D&O-only
    ("board_size",               "Board Size"),
    ("independent_directors",    "Independent Directors"),
]


LOSS_COLUMNS = [
    ("policy_year",  "Policy Year"),
    ("claims",       "Claims"),
    ("incurred",     "Incurred"),
    ("paid",         "Paid"),
    ("reserves",     "Reserves"),
]


# ── Styling ──────────────────────────────────────────────────
# Kept dead-simple so the output renders the same in Excel,
# Numbers, LibreOffice, and Google Sheets.

HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="2F5496")  # readable navy
LABEL_FONT   = Font(bold=True)
TITLE_FONT   = Font(bold=True, size=12)


def _short_id(uuid_str: str) -> str:
    """First block of the UUID — `00000001-0001-…` → `00000001`."""
    return uuid_str.split("-", 1)[0]


def _build_field_tab(wb: Workbook, sub: dict) -> None:
    """One tab per submission, key/value layout.

    Tab name: short_id (e.g. 00000001). Excel limits tab names
    to 31 chars; short_id is 8 — fits with room to spare.
    """
    ws = wb.create_sheet(title=_short_id(sub["id"]))

    # Title row gives the human reading the tab a quick hook.
    title = (
        f"{sub['named_insured']} ({sub['lob']}) — "
        f"{'SHOWCASE' if sub['id'] in SHOWCASE_IDS else 'workflow row'}"
    )
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    # Header row.
    ws["A2"] = "Field"
    ws["B2"] = "Value"
    for cell in (ws["A2"], ws["B2"]):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    # One row per field. Empty values render as blank cells so
    # the operator can tell at a glance what's missing from the
    # seed (e.g. board_size for a GL submission).
    row = 3
    for key, label in FIELD_ORDER:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        if key == "short_id":
            ws.cell(row=row, column=2, value=_short_id(sub["id"]))
        elif key == "showcase":
            ws.cell(
                row=row, column=2,
                value="yes" if sub["id"] in SHOWCASE_IDS else "no",
            )
        else:
            ws.cell(row=row, column=2, value=sub.get(key))
        row += 1

    # Column widths tuned by eye for full UUIDs and long
    # industry descriptions.
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 48


def _build_loss_tab(wb: Workbook, sub: dict) -> None:
    """One tab per submission, loss-history table.

    Tab name: LH-<short_id>. The LH- prefix groups all loss
    tabs together at the end of the workbook so the field tabs
    stay contiguous up front.
    """
    losses = sub.get("loss_history", [])
    ws = wb.create_sheet(title=f"LH-{_short_id(sub['id'])}")

    # Title row mirrors the field tab so each tab stands alone.
    ws["A1"] = (
        f"Loss History — {sub['named_insured']} "
        f"({sub['lob']}, {_short_id(sub['id'])})"
    )
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(LOSS_COLUMNS))}1")

    # Header row.
    for col_idx, (_, header) in enumerate(LOSS_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    # One row per loss-history year. If a submission has no
    # loss history (none currently — every seed has 3 years —
    # but defensive), the table is empty below the header.
    for r_idx, loss in enumerate(losses, start=3):
        for c_idx, (key, _) in enumerate(LOSS_COLUMNS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=loss.get(key))

    # Column widths.
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(LOSS_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def main() -> None:
    wb = Workbook()
    # Drop the auto-created default sheet — every tab we create
    # is intentional; we don't want a stray 'Sheet' tab.
    default = wb.active
    wb.remove(default)

    # Field tabs first (in seed-list order — matches the SUBMISSIONS
    # array, which is the order the operator already knows).
    for sub in SUBMISSIONS:
        _build_field_tab(wb, sub)

    # Loss tabs second so they all cluster at the right end of
    # the tab strip.
    for sub in SUBMISSIONS:
        _build_loss_tab(wb, sub)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(
        f"  + wrote {OUT_PATH}  "
        f"({len(SUBMISSIONS)} field tabs + {len(SUBMISSIONS)} loss tabs)"
    )


if __name__ == "__main__":
    main()
